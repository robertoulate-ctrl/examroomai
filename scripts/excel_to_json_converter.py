#!/usr/bin/env python3
"""
ExamRoom.AI Excel to JSON Converter
Converts Startup Questions.xlsx into structured JSON format for test rule configuration
"""

import json
import openpyxl
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

def load_excel_file(filepath: str) -> openpyxl.Workbook:
    """Load the Excel file"""
    try:
        wb = openpyxl.load_workbook(filepath)
        return wb
    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found")
        raise
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        raise

def extract_startup_questions(sheet) -> List[Dict[str, Any]]:
    """Extract startup questions from sheet (special format)"""
    questions = []
    current_question = None
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        col_a, col_b, col_c, col_d = row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None, row[2] if len(row) > 2 else None, row[3] if len(row) > 3 else None
        
        # Skip empty rows
        if col_a is None and col_c is None:
            continue
        
        # If col_c has data, it's an answer
        if col_c is not None:
            if current_question:
                current_question['answer'] = col_c
                current_question['comments'] = col_d
        # If col_a has data, it's a question
        elif col_a is not None:
            if current_question:
                questions.append(current_question)
            current_question = {
                'row_number': row_idx,
                'question': col_a,
                'answer': None,
                'comments': col_d
            }
    
    # Add the last question
    if current_question:
        questions.append(current_question)
    
    return questions

def convert_excel_to_json(input_file: str, output_file: Optional[str] = None) -> Dict[str, Any]:
    """
    Convert Excel file to JSON format
    Returns the complete JSON structure
    """
    
    if output_file is None:
        output_file = input_file.replace('.xlsx', '.json')
    
    # Load workbook
    wb = load_excel_file(input_file)
    
    # Create main structure
    result = {
        "metadata": {
            "source_file": Path(input_file).name,
            "conversion_date": datetime.now().isoformat(),
            "converter_version": "1.0",
            "description": "ExamRoom.AI Startup Questions - Test Rules Configuration"
        },
        "test_configuration": {
            "organization_info": {},
            "contact_info": {},
            "platform_config": {},
            "exam_settings": {},
            "proctoring_config": {},
            "other_settings": {}
        }
    }
    
    # Process Startup Questions sheet
    print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    
    if 'EX.AI Startup Questions' in wb.sheetnames:
        print("\nProcessing: EX.AI Startup Questions")
        sheet = wb['EX.AI Startup Questions']
        
        questions = extract_startup_questions(sheet)
        
        # Categorize questions
        for q in questions:
            question_text = q['question'].lower() if q['question'] else ""
            
            if "organization" in question_text or "company" in question_text:
                result["test_configuration"]["organization_info"][q['question']] = {
                    "answer": q['answer'],
                    "comments": q['comments']
                }
            elif "contact" in question_text or "reach out" in question_text:
                result["test_configuration"]["contact_info"][q['question']] = {
                    "answer": q['answer'],
                    "comments": q['comments']
                }
            elif "platform" in question_text or "url" in question_text or "console" in question_text:
                result["test_configuration"]["platform_config"][q['question']] = {
                    "answer": q['answer'],
                    "comments": q['comments']
                }
            elif "timer" in question_text or "disconnected" in question_text or "answers" in question_text:
                result["test_configuration"]["exam_settings"][q['question']] = {
                    "answer": q['answer'],
                    "comments": q['comments']
                }
            elif "proctoring" in question_text or "proctor" in question_text or "pay" in question_text:
                result["test_configuration"]["proctoring_config"][q['question']] = {
                    "answer": q['answer'],
                    "comments": q['comments']
                }
            else:
                result["test_configuration"]["other_settings"][q['question']] = {
                    "answer": q['answer'],
                    "comments": q['comments']
                }
        
        print(f"  ✓ Extracted {len(questions)} questions")
    
    if 'Additional Exams Info' in wb.sheetnames:
        print("\nProcessing: Additional Exams Info")
        sheet = wb['Additional Exams Info']
        additional_data = extract_startup_questions(sheet)
        result["additional_exams_info"] = additional_data
        print(f"  ✓ Extracted {len(additional_data)} additional items")
    
    # Save to JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Successfully converted to: {output_file}")
    
    return result

def analyze_excel_structure(input_file: str) -> None:
    """Analyze and print the structure of the Excel file"""
    
    wb = load_excel_file(input_file)
    
    print("\n" + "="*80)
    print("EXCEL FILE STRUCTURE ANALYSIS")
    print("="*80)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n📄 Sheet: {sheet_name}")
        print("-" * 80)
        print(f"   Dimensions: {sheet.dimensions}")
        
        # Show first 15 rows
        print(f"\n   First 15 rows:")
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            non_empty = [v for v in row if v is not None]
            if non_empty:
                print(f"     Row {row_idx:2d}: {row}")
    
    print("\n" + "="*80)

if __name__ == "__main__":
    input_file = "Startup Questions.xlsx"
    output_file = "startup_questions_config.json"
    
    # Analyze structure first
    analyze_excel_structure(input_file)
    
    # Convert to JSON
    result = convert_excel_to_json(input_file, output_file)
    
    # Print summary
    print(f"\n📊 Conversion Summary:")
    print(f"   Organization Info: {len(result['test_configuration']['organization_info'])} items")
    print(f"   Contact Info: {len(result['test_configuration']['contact_info'])} items")
    print(f"   Platform Config: {len(result['test_configuration']['platform_config'])} items")
    print(f"   Exam Settings: {len(result['test_configuration']['exam_settings'])} items")
    print(f"   Proctoring Config: {len(result['test_configuration']['proctoring_config'])} items")
    print(f"   Other Settings: {len(result['test_configuration']['other_settings'])} items")
